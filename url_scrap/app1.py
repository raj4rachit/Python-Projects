import requests
from bs4 import BeautifulSoup
import urllib.parse  # For handling relative URLs
import openpyxl  # For creating Excel file

def download_urls_ethically(url, excel_file_path, delay=2):  # Add delay parameter
    """Downloads URLs from the given website (with ethical considerations) and saves them to an Excel file.

    Args:
        url (str): The URL of the website to scrape.
        excel_file_path (str): The path to save the Excel file.
        delay (float, optional): The delay in seconds between requests (default: 2).
    """

    try:
        response = requests.get(url)
        response.raise_for_status()  # Raise an exception for non-200 status codes

        soup = BeautifulSoup(response.content, 'html.parser')

        # Extract all links (anchor tags) from the website
        links = soup.find_all('a')

        # Create an Excel workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1).value = "URL"  # Header for the first column

        # Iterate through the links, handle relative URLs, and save to Excel (with delay)
        row_count = 2
        for link in links:
            href = link.get('href')
            if href:
                # Handle relative URLs (starting with '/') or empty hrefs
                if href.startswith('/'):
                    href = urllib.parse.urljoin(url, href)  # Join relative URL with base URL
                elif not href.startswith('http'):
                    continue  # Skip empty or non-http(s) links

                ws.cell(row=row_count, column=1).value = href
                row_count += 1

                # Add a delay between requests to be respectful
                import time
                time.sleep(delay)

        wb.save(excel_file_path)
        print(f"Downloaded URLs from {url} (with delays) and saved to {excel_file_path}")

    except requests.exceptions.RequestException as e:
        print(f"Error downloading URLs: {e}")

# Example usage with user confirmation
if __name__ == "__main__":
    website_url = "https://yojonesy.com/"
    excel_file_path = "downloaded_urls.xlsx"

    confirmation = input(f"Downloading URLs from {website_url} may violate their terms of service. Are you sure you want to proceed (y/n)? ")
    if confirmation.lower() == 'y':
        download_urls_ethically(website_url, excel_file_path)
    else:
        print("Aborting download.")
